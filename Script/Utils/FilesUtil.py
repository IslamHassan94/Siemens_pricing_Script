import datetime
import os

import openpyxl
import win32com.client as win32

from Script.Config import Config_Setup


def protect_excel_with_password(file_path: str, password: str):
    # Check if the file exists
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"The file '{file_path}' does not exist.")

    # Create an instance of Excel
    excel = win32.Dispatch("Excel.Application")

    # Open the Excel file
    workbook = excel.Workbooks.Open(file_path)

    # Set the password to protect the file
    workbook.Password = password

    # Save the workbook with the password
    workbook.Save()

    # Close the workbook and quit Excel
    workbook.Close()
    excel.Quit()

    print(f"The file '{file_path}' has been password protected.")


def clean_dir(directory):
    for filename in os.listdir(directory):
        file_path = os.path.join(directory, filename)
        try:
            if os.path.isfile(file_path) or os.path.islink(file_path):
                os.unlink(file_path)  # Remove file or link
            elif os.path.isdir(file_path):
                shutil.rmtree(file_path)  # Remove directory and its contents
        except Exception as e:
            print(f"Failed to delete {file_path}. Reason: {e}")


def get_first_file_name():
    files_list = []
    for filename in os.listdir(downloads_path):
        file_path = os.path.join(downloads_path, filename)
        print(file_path)
        files_list.append(filename)
        print(filename)
    return files_list[0]


def get_files_list(path):
    files_list = []
    for filename in os.listdir(path):
        file_path = os.path.join(path, filename)
        files_list.append(filename)
    return files_list


def take_backup(source_file_path, destination_dir):
    today_date = datetime.datetime.today().strftime('%d-%m-%Y')
    file_name, file_extension = os.path.splitext(os.path.basename(source_file_path))
    new_file_name = f"{file_name}_{today_date}{file_extension}"
    new_file_path = os.path.join(destination_dir, new_file_name)
    shutil.copy2(source_file_path, new_file_path)
    return new_file_path


def get_file_fullName(partial_file_name):
    sheets = get_files_list(Config_Setup.input_sheets_dir)
    for sheet in sheets:
        if partial_file_name in sheet:
            return sheet


def get_file_fullName_with_path(partial_file_name, path):
    sheets = get_files_list(path)
    for sheet in sheets:
        if partial_file_name in sheet:
            return sheet


# ReusableAsset_COE00079
def get_file_fullName_by_keyword_in_name(partial_file_name, keyword):
    sheets = get_files_list(Config_Setup.input_sheets_dir)
    for sheet in sheets:
        if partial_file_name in sheet and keyword in sheet:
            return sheet


def get_file_fullName_by_keyword_not_in_name(partial_file_name, keyword):
    sheets = get_files_list(Config_Setup.input_sheets_dir)
    for sheet in sheets:
        if partial_file_name in sheet and keyword not in sheet:
            return sheet


def is_valid_excel_date(serial_date):
    # Excel uses 1-based serial dates from 1/1/1900 (serial number 1) to 12/31/9999
    return 1 <= serial_date <= 2958465  # 2958465 corresponds to 12/31/9999


def clear_filters_and_unhide_rows(file_path, sheet_name=None):
    # Load the workbook and select the sheet
    wb = openpyxl.load_workbook(file_path)

    # Select the active sheet or the specific sheet if provided
    sheet = wb.active if sheet_name is None else wb[sheet_name]

    # Clear any filters
    sheet.auto_filter = None

    # Unhide all rows
    for row in sheet.iter_rows():
        sheet.row_dimensions[row[0].row].hidden = False

    # Save the workbook after modification
    wb.save(file_path)


def create_empty_excel(file_path):
    # Ensure the directory exists
    os.makedirs(os.path.dirname(file_path), exist_ok=True)

    # Create a new empty Excel workbook
    workbook = openpyxl.Workbook()

    # Save the workbook at the specified path
    workbook.save(file_path)
